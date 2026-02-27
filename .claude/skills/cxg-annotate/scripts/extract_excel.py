#!/usr/bin/env python3
"""Extract supplementary Excel file to plain text (TSV).

Usage:
    uv run python scripts/extract_excel.py <file.xlsx> [sheet_index]

Output: tab-separated text to stdout, one sheet per section separated by
a header line: "=== Sheet: <name> ==="

Requires: openpyxl (included in uv environment via oaklib extras or add
with `uv add openpyxl`)
"""
from __future__ import annotations

import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run: uv add openpyxl", file=sys.stderr)
    sys.exit(1)


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: extract_excel.py <file.xlsx> [sheet_index]", file=sys.stderr)
        sys.exit(1)

    path = sys.argv[1]
    sheet_filter = int(sys.argv[2]) if len(sys.argv) > 2 else None

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    for idx, sheet_name in enumerate(wb.sheetnames):
        if sheet_filter is not None and idx != sheet_filter:
            continue

        ws = wb[sheet_name]
        print(f"=== Sheet: {sheet_name} ===")

        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            print("\t".join(cells))

        print()


if __name__ == "__main__":
    main()
